"""
Elective Service — Dynamic term-based elective management.
Supports upload from: Excel, PDF, Image (OCR via GPT-4o), or plain text.
Data is stored in system_context.yaml and refreshed each term.
LLM-powered query answering with language detection.
"""

import os
import base64
import logging
import yaml
import pdfplumber
import openpyxl
from typing import List, Dict, Union, Any

from dotenv import load_dotenv
from langchain_openai import ChatOpenAI
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.messages import HumanMessage
from langsmith import traceable

from advisor_ai.constants import ELECTIVE_QUERY_PROMPT

load_dotenv()

# Configure logging
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

# Path to the system context YAML
CONTEXT_FILE = os.path.join(os.path.dirname(__file__), "system_context.yaml")


class ElectiveService:
    """Manages elective courses via YAML. Supports multi-format input, LLM queries."""

    def __init__(self):
        self.context_file = CONTEXT_FILE
        self._ensure_context_file()

        self.llm = ChatOpenAI(
            model=os.getenv("OPENAI_LLM_MODEL", "gpt-4o-mini"),
            temperature=0.3,
        )

        # Vision model for image OCR
        self.vision_llm = ChatOpenAI(
            model="gpt-4o-mini",
            temperature=0,
            max_tokens=2000,
        )

        logger.info("Elective Service initialized (LLM-powered)")

    def _ensure_context_file(self):
        """Create default context file if it doesn't exist."""
        if not os.path.exists(self.context_file):
            default = {
                "active_term": "Spring-2026",
                "electives": [],
            }
            self._save_context(default)

    def _load_context(self) -> dict:
        """Load the system context from YAML."""
        try:
            with open(self.context_file, "r", encoding="utf-8") as f:
                return yaml.safe_load(f) or {}
        except Exception as e:
            logger.error(f"Error loading context: {e}")
            return {}

    def _save_context(self, context: dict):
        """Save the system context to YAML."""
        try:
            with open(self.context_file, "w", encoding="utf-8") as f:
                yaml.dump(context, f, default_flow_style=False, allow_unicode=True)
        except Exception as e:
            logger.error(f"Error saving context: {e}")

    # ── Getters / Setters ────────────────────────────────────────────

    def get_active_term(self) -> str:
        """Get the current active term."""
        context = self._load_context()
        return context.get("active_term", "Unknown")

    def get_electives(self) -> List[Union[str, Dict[str, Any]]]:
        """Get the list of available electives for the active term."""
        context = self._load_context()
        return context.get("electives", [])

    def set_term(self, term: str):
        """Update the active term."""
        context = self._load_context()
        context["active_term"] = term
        self._save_context(context)
        logger.info(f"Active term updated to: {term}")

    def set_electives(self, electives: list):
        """Set the elective list (can be strings or dicts with details)."""
        context = self._load_context()
        context["electives"] = electives
        self._save_context(context)
        logger.info(f"Electives updated: {len(electives)} courses")

    # ── Upload Methods (Multi-Format) ────────────────────────────────

    @traceable(name="Upload Electives", run_type="chain")
    def upload(self, source: str) -> list:
        """
        Auto-detect the format and upload electives.
        source: file path (Excel, PDF, image) OR raw text.
        """
        if os.path.isfile(source):
            ext = os.path.splitext(source)[1].lower()
            if ext in (".xlsx", ".xls"):
                return self.upload_from_excel(source)
            elif ext == ".pdf":
                return self.upload_from_pdf(source)
            elif ext in (".png", ".jpg", ".jpeg", ".webp", ".bmp"):
                return self.upload_from_image(source)
            else:
                # Try reading as text file
                try:
                    with open(source, "r", encoding="utf-8") as f:
                        return self.upload_from_text(f.read())
                except Exception as e:
                    logger.error(f"Error reading text file: {e}")
                    return []
        else:
            # Treat as raw text input
            return self.upload_from_text(source)

    def upload_from_excel(self, file_path: str) -> list:
        """Parse electives from Excel."""
        electives = []
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            if ws is None:
                raise ValueError("Excel file has no active sheet")

            # Read header row to detect columns
            headers = [str(cell.value or "").strip().lower() for cell in ws[1]]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue

                if len(headers) <= 1:
                    electives.append(str(row[0]).strip())
                else:
                    record = self._parse_excel_row(row, headers)
                    electives.append(record) # type: ignore

            self.set_electives(electives)
            return electives
        except Exception as e:
            logger.error(f"Error reading Excel: {e}")
            raise

    def _parse_excel_row(self, row: tuple, headers: list) -> dict:
        """Helper to parse a single Excel row based on headers."""
        record = {}
        for i, header in enumerate(headers):
            if i < len(row) and row[i]:
                val = str(row[i]).strip()
                if "code" in header: record["code"] = val
                elif any(k in header for k in ["name", "course", "مادة"]): record["name"] = val
                elif any(k in header for k in ["instructor", "doctor", "دكتور"]): record["instructor"] = val
                elif any(k in header for k in ["day", "يوم"]): record["day"] = val
                elif any(k in header for k in ["time", "وقت"]): record["time"] = val
                elif any(k in header for k in ["credit", "hour", "ساعة"]): record["credits"] = val
        
        if "name" not in record and row:
             record["name"] = str(row[0]).strip()
        return record

    def upload_from_pdf(self, file_path: str) -> list:
        """Parse electives from a PDF."""
        electives = []
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    # Try table extraction first
                    tables = page.extract_tables()
                    if tables:
                        for table in tables:
                            for row in table[1:]:  # Skip header
                                if row and row[0] and str(row[0]).strip():
                                    if len(row) >= 2:
                                        record = {"name": str(row[0]).strip()}
                                        if len(row) >= 2 and row[1]: record["code"] = str(row[1]).strip()
                                        if len(row) >= 3 and row[2]: record["instructor"] = str(row[2]).strip()
                                        if len(row) >= 4 and row[3]: record["time"] = str(row[3]).strip()
                                        electives.append(record)
                                    else:
                                        electives.append(str(row[0]).strip()) # type: ignore
                    else:
                        # Fallback: line-based
                        text = page.extract_text()
                        if text:
                            for line in text.strip().split("\n"):
                                line = line.strip()
                                if line and len(line) > 2 and not line.startswith("#"):
                                    electives.append(line) # type: ignore

            self.set_electives(electives)
            return electives
        except Exception as e:
            logger.error(f"Error reading PDF: {e}")
            raise

    def upload_from_image(self, file_path: str) -> list:
        """Extract elective data from an image using GPT-4o vision OCR."""
        try:
            with open(file_path, "rb") as f:
                image_data = base64.b64encode(f.read()).decode("utf-8")

            ext = os.path.splitext(file_path)[1].lower()
            mime_type = f"image/{ext[1:]}" if ext != ".jpg" else "image/jpeg"

            message = HumanMessage(content=[
                {"type": "text", "text": (
                    "Extract ALL course/elective names from this image. "
                    "Return ONLY a YAML list. For each course include visible details. "
                    "Format:\n- name: Course Name\n  code: CODE\n  credits: 3\n"
                    "Return ONLY the YAML."
                )},
                {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{image_data}"}},
            ])

            result = self.vision_llm.invoke([message])
            yaml_text = result.content.strip()

            # Clean markdown code fences
            if "```" in yaml_text:
                yaml_text = yaml_text.replace("```yaml", "").replace("```", "").strip()

            electives = yaml.safe_load(yaml_text)
            if not isinstance(electives, list):
                electives = [electives]

            self.set_electives(electives)
            logger.info(f"Extracted {len(electives)} courses from image")
            return electives
        except Exception as e:
            logger.error(f"Error reading image: {e}")
            raise

    def upload_from_text(self, text: str) -> list:
        """Parse electives from raw text."""
        electives = []
        # Handle comma-separated if single line
        if "," in text and "\n" not in text:
            items = text.split(",")
        else:
            items = text.strip().split("\n")

        for item in items:
            item = item.strip().strip("-").strip("•").strip("*").strip()
            if item and len(item) > 1:
                electives.append(item)

        self.set_electives(electives)
        return electives

    # ── Query Methods ────────────────────────────────────────────────

    @traceable(name="Elective Query", run_type="chain")
    def query(self, question: str) -> str:
        """LLM-powered elective query."""
        try:
            term_data = self._build_term_context()

            chain = ChatPromptTemplate.from_messages([
                ("system", ELECTIVE_QUERY_PROMPT),
                ("human", "{question}"),
            ]) | self.llm

            result = chain.invoke({
                "term_data": term_data,
                "question": question,
            })
            return result.content
        except Exception as e:
            logger.error(f"LLM query error, falling back: {e}")
            return self.get_electives_text()

    def _build_term_context(self) -> str:
        """Build a text summary of current term electives."""
        term = self.get_active_term()
        electives = self.get_electives()

        if not electives:
            return f"Term: {term}\nNo electives available yet for this term."

        lines = [f"Term: {term}", f"Available Electives ({len(electives)}):", ""]
        for i, e in enumerate(electives, 1):
            if isinstance(e, dict):
                parts = [f"{i}."]
                if e.get("code"): parts.append(f"[{e['code']}]")
                parts.append(e.get("name", "Unknown"))
                if e.get("credits"): parts.append(f"({e['credits']} credits)")
                lines.append(" ".join(parts))
            else:
                lines.append(f"{i}. {e}")

        return "\n".join(lines)

    def get_electives_text(self) -> str:
        """Get a formatted string of available electives (fallback)."""
        return self._build_term_context()
